# coding: utf-8
import sys
import os
import csv
import glob
import argparse

import openpyxl as px

def main():
    parser = argparse.ArgumentParser(description='Excel-CSV間の読み書き')

    parser.add_argument('arg1', help='e2c or c2e を指定')
    parser.add_argument('arg2', help='e2c: エクセルファイルを指定, c2e: ディレクトリを指定')

    args = parser.parse_args()

    if args.arg1 == 'e2c':
        convert_book_to_csvs(args.arg2)
    elif args.arg1 == 'c2e':
        convert_csvs_to_book(args.arg2)

    print ('FINISHED')

def convert_book_to_csvs(excelfile):
    """
    Excelファイルの各シートをcsvへ変換する

    Parameters
    ----------
    excelfile : string
        変換対象のExcelファイルパス
    """

    abspath = os.path.abspath(excelfile)

    if not os.path.exists(abspath):
        print ('[ERROR] EXCEL FILE: ', excelfile, ' is not found')
        sys.exit()

    print ('EXCELFILE : ', abspath)

    convert_book_to_csvs_core(abspath)

def convert_book_to_csvs_core(exfilepath):
    """
    Excelファイルの各シートをcsvへ変換する

    Parameters
    ----------
    excelfile : string
        変換対象のExcelファイルパス
    """

    workbook = px.load_workbook(exfilepath, read_only=True, keep_vba=False)
    savedir = os.path.dirname(exfilepath)

    for sheetname in workbook.sheetnames:
        worksheet = workbook[sheetname]

        if not worksheet.max_row - 1:
            continue

        csv_filepath = os.path.join(savedir, str(sheetname) + '.csv')
        convert_sheet_to_csv(worksheet, csv_filepath)
        print ('CSVFILE : ', csv_filepath)

def convert_sheet_to_csv(worksheet, filepath):
    """
    Excelファイルのシートをcsvへ変換する

    Parameters
    ----------
    worksheet : 
        ワークシート
    filepath : string
        保存ファイルパス
    """
    rows = enumerate_read_rows_from_sheet(worksheet)
    write_csv(rows, filepath, 'utf-8', ',')

def convert_csvs_to_book(exceldir):
    excelfiles = glob.glob(exceldir + '/*.xlsx')

    if not excelfiles:
        print ('[ERROR] DIRECTORY: ', exceldir, ' is not in excelfile')
        sys.exit()

    csvfiles = glob.glob(exceldir + '/*.csv')

    if not csvfiles:
        print ('[ERROR] DIRECTORY: ', exceldir, ' is not in csvfiles')
        sys.exit()

    abspath = os.path.abspath(excelfiles[0])
    convert_csvs_to_book_core(abspath, csvfiles)

def convert_csvs_to_book_core(excelfile, csvfiles):
    """
    csvファイルリストをExcelファイルへ変換する
    """
    workbook = px.load_workbook(excelfile)
    print ('EXCELFILE : ', excelfile)

    for csvfile in csvfiles:
        [sheetname, ext] = os.path.splitext(os.path.basename(csvfile))

        if contains_sheet(workbook, sheetname):
            worksheet = workbook[sheetname]
            clear_sheet(worksheet)
        else:
            worksheet = workbook.create_sheet(title=sheetname)

        print ('EXCELFILE : ', csvfile)

        rows = enumerate_read_csv(csvfile, 'utf-8', ',')
        write_rows_to_sheet(worksheet, rows)

    workbook.save(excelfile)

def clear_sheet(worksheet):
    for row in worksheet.iter_rows():
        for cell in row:
            cell.value = None

def enumerate_read_rows_from_sheet(worksheet):
    for cols in worksheet.rows:
        yield [str(col.value or '') for col in cols]

def read_rows_from_sheet(worksheet):
    return [row for row in enumerate_read_rows_from_sheet(worksheet)]

def write_rows_to_sheet(worksheet, rows, start_row=1, start_col=1):
    for y, row in enumerate(rows):
        for x, cell in enumerate(row):
            worksheet.cell(row=start_row+y, column=start_col+x, value=cell)

def enumerate_read_csv(filepath, encoding, delimiter):
    with open(filepath, encoding=encoding) as fp:
        reader = csv.reader(fp, delimiter=delimiter)
        for row in reader:
            yield row

def read_csv(filepath, encoding, delimiter):
    return [row for row in enumerate_read_csv(filepaath, encoding, delimiter)]

def write_csv(rows, filepath, encoding, delimiter):
    with open(filepath, 'w', newline='', encoding=encoding) as fp:
        writer = csv.writer(fp, lineterminator='\n')
        for row in rows:
            writer.writerow(row)

def contains_sheet(workbook, sheetname):
    for sh in workbook.sheetnames:
        if sh == sheetname:
            return True
    return False

if __name__ == '__main__':
    main()